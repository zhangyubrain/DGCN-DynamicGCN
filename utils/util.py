from scipy import stats
import matplotlib.pyplot as plt
import numpy as np
import torch
from scipy.io import loadmat
from sklearn.model_selection import StratifiedKFold
from sklearn.model_selection import LeaveOneOut 
import numpy as np
import h5py
import pandas as pd
from sklearn.model_selection import KFold
import csv
from os import listdir
import scipy.io as sio
import os.path as osp
from sklearn.metrics import roc_curve, confusion_matrix, auc
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns
import torch.nn.functional as F
import re
import os
import random

def normal_transform_train(x):
    #xt, lamb = stats.boxcox(x - torch.min(x) + 1)
    lamb = 0
    #xt_torch = xt
    xt_mean = torch.mean(x).float()
    xt_std = torch.std(x).float()
    xt_norm = (x-xt_mean)/xt_std
    return xt_norm,lamb,xt_mean, xt_std


def normal_transform_test(x,lamb, xt_mean, xt_std):
    res = (x-xt_mean)/xt_std
    return res

def count_parameters(model):
    for p in model.parameters():
        if p.requires_grad:
            print('Number of parameter', p.numel())
    #return sum(p.numel() for p in model.parameters() if p.requires_grad)

def train_val_test_split( labels, numbers, kfold = 10, rep=1):

    y_ind = np.array(labels)
    x_ind = range(numbers)

    skf = StratifiedKFold(n_splits=kfold, shuffle=True, random_state=8)
    # skf2 = StratifiedKFold(n_splits=kfold, shuffle=True, random_state = 666)
    test_index = list()
    train_index = list()
    validation_index = list()

    for a, b in skf.split(x_ind, y_ind):
        test_index.append(b)
        #######如果多划分验证集用下代码：
        # temp1, temp2 = list(skf2.split(a, y_ind[a]))[0]
        # c = a[temp1]
        # d = a[temp2]
        # train_index.append(c)
        # validation_index.append(d)      
        train_index.append(a)
        validation_index.append(b)    
        
    # raw_index = list(skf.split(x_ind, y_ind))
    # raw_train, raw_test = raw_index[0]
    # i = 0
    # for a, b in skf2.split(raw_test, y_ind[raw_test]):
    #     test_index.append(raw_test[b])     
    #     train_index.append(np.sort(np.concatenate((raw_train,raw_test[a]))))
    #     # train_index.append(raw_train)
    #     validation_index.append(raw_test[b])
    #     i =i+1

    tr_f=[]
    te_f=[]
    val_f=[]
    for i in range(kfold):
        train_id = train_index[i]
        test_id = test_index[i]
        val_id = validation_index[i]
        
        if rep >1:
            for k in range(rep):
                if k < 1:                    
                    tr_index = train_id
                else:
                    tr_index = np.append(tr_index,train_id+k*numbers)

        else:
            tr_index = train_id

        tr_f.append(tr_index)
        te_f.append(test_id)
        val_f.append(val_id)
    return tr_f,te_f,val_f

def get_tridx(direction, filename):
    root = direction
    # rule = re.compile('^[a-zA-z]{1}.*$')
    all_result=[]
    direction = os.path.join(root, filename)
    f = open(direction,'r')
    lines = f.readlines()[1:32]
    for line in lines:
        # if rule.match(line[1]) is None:
        new_line = re.findall('\d+',line)
        all_result.extend(new_line)
    
    number_array = np.array(all_result).astype(int)
    return number_array

def get_tridx2(direction, filename):
    root = direction
    # rule = re.compile('^[a-zA-z]{1}.*$')
    all_result=[]
    direction = os.path.join(root, filename)
    f = open(direction,'r')
    lines = f.readlines()[1]
    str1 = lines.split('[') # 用逗号分割str字符串，并保存到列表
    str2 = str1[-1].split(']')[0] # 用逗号分割str字符串，并保存到列表
    s_f = str2.split(',')
    number_array = np.array(s_f).astype(int)
    return number_array

def get_index(all_file, labels, pcds, numbers):
    y = np.array(labels)
    all_idx = set(range(numbers))
    
    tr_idx = []
    te_idx = []
    

    for j in range(len(all_file)):

        idx = get_tridx(all_file[j][0], all_file[j][1])
        idx = set(idx)
        te = all_idx - idx   
        idx = list(idx)
        te = list(te)
        tr_idx.append(idx)
        te_idx.append(te)      

    val_idx = te_idx
        
    return tr_idx, val_idx, te_idx, all_file

def split_reproduce(labels, pcd, kfold = 10, number=666):

    y_ind = np.array(labels)
    x_ind = range(len(labels))

    skf = StratifiedKFold(n_splits=kfold, shuffle=True, random_state=number)
    # skf2 = StratifiedKFold(n_splits=kfold-1, shuffle=True, random_state = 2)

    tr_f=[]
    te_f=[]
    for train, test in skf.split(x_ind, y_ind):

        tr_f.append(train)
        te_f.append(test)
        
    return tr_f,te_f

def site_split(labels, pcd, kfold = 10, number=666):

    site = [info[0] for info in pcd]
    site_array = np.array(site)
    elements = np.unique(site_array)
    # all_idx, all_y =[], []
    y_ind = np.array(labels)

    skf = StratifiedKFold(n_splits=kfold, shuffle=True, random_state=number)
    # skf2 = StratifiedKFold(n_splits=kfold-1, shuffle=True, random_state = 2)

    tr_f=[]
    te_f=[]
    val_f=[]
    for e in elements:
        idx = np.where(site_array==e)[0]
        tem_y = y_ind[idx]
        # y_0 = np.where(tem_y==0)[0]
        # y_1 = np.where(tem_y==1)[0]
        # if len(y_0)>len(y_1):
        #     y = y_0[:len(y_1)]
        #     idx_sift = np.concatenate((y_1, y))
        # else:
        #     y = y_1[:len(y_0)]
        #     idx_sift = np.concatenate((y_0, y))
        # all_idx.append(idx)
        # y_new = tem_y[idx_sift]
        # all_y.append(y_new)
        test_index = list()
        train_index = list()
        validation_index = list()
        
        for a, b in skf.split(idx, tem_y):
            
            train = idx[a]
            # test = idx[idx_sift[b]]
            test = idx[b]
            # np.random.shuffle(train)
            # np.random.shuffle(test)

            test_index.append(test)
            # tr, val = list(skf2.split(a, idx_sift[a]))[0]
            # c = idx[a][tr]
            # d = idx[a][val]
            # train_index.append(c)
            # validation_index.append(d) 
            
            train_index.append(train)
            validation_index.append(test) 
        tr_f.append(train_index)
        val_f.append(validation_index)
        te_f.append(test_index)
    tr_f = np.array(tr_f)   
    val_f = np.array(val_f)   
    te_f = np.array(te_f)   
     
    tr, val, te = [],[], []
    for i in range(kfold):
        # tr_tem = np.concatenate((tr_f[0,i], tr_f[1,i], tr_f[2,i], tr_f[3,i]), axis=-1)
        # val_tem = np.concatenate((val_f[0,i], val_f[1,i], val_f[2,i], val_f[3,i]), axis=-1)
        # te_tem = np.concatenate((te_f[0,i], te_f[1,i], te_f[2,i], te_f[3,i]), axis=-1)   
        tr_tem = np.concatenate((tr_f[0,i], tr_f[1,i], tr_f[2,i]), axis=-1)
        val_tem = np.concatenate((val_f[0,i], val_f[1,i], val_f[2,i]), axis=-1)
        te_tem = np.concatenate((te_f[0,i], te_f[1,i], te_f[2,i]), axis=-1)
        tr_tem.sort()
        val_tem.sort()
        te_tem.sort()
        tr.append(tr_tem)
        val.append(val_tem)
        te.append(te_tem)
        
    return tr,val,te

def train_val_split_LO(y_ind, numbers):

    x_ind = range(numbers)
    loo = LeaveOneOut()
    train_index = list()
    validation_index = list()

    for a, b in loo.split(x_ind):
        train_index.append(a)
        validation_index.append(b)    
        
    return train_index , validation_index

def get_EHRinfo(data_dir, varibale):
# Load precomputed fMRI connectivity networks
    onlyfile = [f for f in listdir(data_dir) if osp.isfile(osp.join(data_dir, f))]
    onlyfiles = [data_dir+'/'+f for f in onlyfile]
    site_array = np.zeros((len(onlyfiles)))
    for i,file_name in enumerate(onlyfiles):
        data = sio.loadmat(file_name)
        site_array[i] = int(data[varibale])
  
    site_torch = torch.from_numpy(site_array).long()  # classification

    return site_torch


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    if isinstance(value[0], tuple):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
        workbook.save(path)
        
    else: 
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for i in range(0, index):
                sheet.cell(row=i+1, column=1, value=str(value[i]))
        workbook.save(path)
    
def plot_fea(title, x_fea, label):
    sns.set()
    plt.figure()
    plt.title('{} feature'.format(title))
    if 'hand' in title:
        sns.scatterplot(x=x_fea[:,0], y=x_fea[:,1], hue=label, palette="ch:s=.25,rot=-.25")
    elif 'site' in title:
        label=label.astype(np.int32)
        sns.scatterplot(x=x_fea[:,0], y=x_fea[:,1], hue=label, palette="Set1")
    else: 
        sns.scatterplot(x=x_fea[:,0], y=x_fea[:,1], hue=label, palette="ch:s=.25,rot=-.25")
    plt.savefig(r'E:\OPT\research\fMRI\PRGNN_fMRI-myModify\result\ACC_LOSS\254_aal_90roi_visual_feature\{}'.format(title))
    plt.show()
    
def plot_distr(title, output, label):
    data1 = output[:,0]
    data2 = output[:,1]
    idx = np.zeros(len(label))+1
    data = {'data1':data1,'data2':data2,'y':label,'idx':idx}
    data = pd.DataFrame(data)
    plt.figure(0)
    ax = sns.violinplot(x='idx', y='data1', hue='y', data=data, palette="Set2", split=True, scale="count", inner=None)
    plt.title('{}'.format(title))
    plt.savefig(r'E:\OPT\research\fMRI\PRGNN_fMRI-myModify\result\ACC_LOSS\254_aal_90roi_visual_feature\{}'.format(title))
    # ax = sns.violinplot(x='idx', y='data2', hue='y', data=data, palette="Set2", split=True, scale="count", inner=None)
    plt.show()
    
    
def plot_ROC(labels, preds, save_file, name):
    fpr, tpr, thresholds_keras = roc_curve(labels, preds)
    Fauc = auc(fpr, tpr)
    print("AUC : ", Fauc)
    plt.figure()
    plt.plot([0, 1], [0, 1], 'k--')
    plt.plot(fpr, tpr, label='area = {:.3f}'.format(Fauc))
    plt.xlabel('False positive rate')
    plt.ylabel('True positive rate')
    plt.title('ROC curve')
    plt.legend(loc='best') 

    plt.savefig(save_file+"/{}_ROC_10x_over.png".format(name))
    plt.show()

def sens_spec(predict, label):
    matrix = confusion_matrix(label, predict)    
    tp = float(matrix[0][0])/np.sum(matrix[0])
    tn = float(matrix[1][1])/np.sum(matrix[1])
    return tp, tn

def js_div(p_output, q_output, get_softmax=True):
    """
    Function that measures JS divergence between target and output logits:
    """
    KLDivLoss = torch.nn.KLDivLoss()
    if get_softmax:
        p_output = F.softmax(p_output)
        q_output = F.softmax(q_output)
    mean_output = (p_output + q_output )/2
    return (KLDivLoss(mean_output.log(), p_output) + KLDivLoss(mean_output.log(), q_output))/2